#!/usr/bin/python
##########
#Author - Shaunak M Deshmukh

#Usage - python script_name.py <your gmail password for account configured at line 28>

#########

import openpyxl,sys,smtplib

wb=openpyxl.load_workbook('/Users/shaunak/Downloads/CompanyOutings3.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')
#lastcol=(str(sheet.max_column())
#print(type(sheet.max_column()))
latestmonth=sheet.cell(row=1, column=3).value
unpaidmembers={}
for r in range(2, 4):
    payment=sheet.cell(row=r, column=3).value
#    print(payment)
    if payment!='paid':
      name=sheet.cell(row=r, column=1).value
      email=sheet.cell(row=r, column=2).value
      unpaidmembers[name]=email

smtpObj=smtplib.SMTP('smtp.gmail.com',587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('shaunak7237@gmail.com', sys.argv[1])

for name,email in unpaidmembers.items():
    body= "Subject : Reminder for your company outing payment due for %s" %latestmonth 
    print("Sending email to %s" %email) 
    sendmailstatus=smtpObj.sendmail('shaunak7237@gmail.com',email,body)  
    if sendmailstatus != {}:
        print ("There was an issue sending email to %s" %email)
smtpObj.quit()  
    
